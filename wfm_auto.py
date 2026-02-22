import requests
import os
import time
from datetime import datetime
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook

# ==============================
# KONFIGURASI
# ==============================
URL = "https://wfm.telkom.co.id/jw/web/json/plugin/org.telkom.co.id.ExportWorkorder/service"
COOKIE = "JSESSIONID=SmxzuCYByN_SD2LuI-Fl0BpcSGKJgY8QpzbRjAHl.bima-dc845d86d-4l8c5; 2583a12ba4a45f8c3321a228b61ba02c=1e8d254bbd32f18743dea220cf346bff"
SPREADSHEET_ID = "17SL1HE0YSyJCcq68znfPOFewUp1ctrHckBKUfHqEdh4"
SHEET_NAME = "WFM"  # <-- ganti nama sheet di sini
LOG_FILE = "wfm_auto.log"

# ==============================
# LOGGING
# ==============================
def log(message):
    print(message)
    with open(LOG_FILE, "a") as f:
        f.write(f"{datetime.now()} - {message}\n")

# ==============================
# RANGE TANGGAL 10–10 OTOMATIS
# ==============================
def get_date_range():
    today = datetime.today()
    start_date = today.replace(day=10)
    if start_date.month == 12:
        end_date = start_date.replace(year=start_date.year+1, month=1)
    else:
        end_date = start_date.replace(month=start_date.month+1)
    return start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")

# ==============================
# DOWNLOAD FILE WFM
# ==============================
def download_file(DATECREATED_FROM, DATECREATED_TO):
    log(f"Mulai download dari WFM ({DATECREATED_FROM} s/d {DATECREATED_TO})")
    headers = {"Content-Type": "application/json", "Cookie": COOKIE}
    payload = {
        "filters": {
            "C_STATUS": "STARTWORK",
            "C_WORKZONE": "PBL,SKP,TGS,LCE,PTN,KRZ,GND,JTO,KKH,LMJ,PII,SDO,TPH,YSN,BNL,GDW,GRA,NJA,PSN,TOS",
            "DATECREATED_FROM": DATECREATED_FROM,
            "DATECREATED_TO": DATECREATED_TO,
            "C_SCORDERNO": "WSA"
        },
        "page": 1,
        "pageSize": 10,
        "SORT": "DESC",
        "ORDER_BY": "datecreated"
    }

    try:
        r = requests.post(URL, json=payload, headers=headers)
        if r.status_code == 200:
            with open("temp.xlsx", "wb") as f:
                f.write(r.content)
            log("Sukses download file WFM")
            return True
        else:
            log(f"Gagal download: {r.status_code} - {r.text}")
            return False
    except Exception as e:
        log(f"Error download: {e}")
        return False

# ==============================
# UPLOAD KE GOOGLE SHEET
# ==============================
def upload_to_google_sheet():
    log(f"Mulai upload ke Google Spreadsheet sheet '{SHEET_NAME}'")
    try:
        scope = ["https://spreadsheets.google.com/feeds",
                 "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_name("credentials.json", scope)
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_key(SPREADSHEET_ID)

        # cek apakah sheet WFM ada
        try:
            sheet = spreadsheet.worksheet(SHEET_NAME)
        except gspread.WorksheetNotFound:
            sheet = spreadsheet.add_worksheet(title=SHEET_NAME, rows="1000", cols="50")

        wb = load_workbook("temp.xlsx", read_only=True)
        ws = wb.active
        data = [list(row) for row in ws.iter_rows(values_only=True)]
        wb.close()

        sheet.clear()
        sheet.update(data)
        os.remove("temp.xlsx")
        log("Sukses upload ke Google Spreadsheet (WFM)")

    except Exception as e:
        log(f"Error upload: {e}")

# ==============================
# LOOP OTOMATIS 15 MENIT
# ==============================
if __name__ == "__main__":
    while True:
        DATECREATED_FROM, DATECREATED_TO = get_date_range()
        if download_file(DATECREATED_FROM, DATECREATED_TO):
            upload_to_google_sheet()
        else:
            log("Proses dibatalkan karena download gagal")
        log("Menunggu 15 menit sebelum proses berikutnya...\n")
        time.sleep(15 * 60)
